"""
excel_manager.py — 엑셀 파일 읽기/쓰기/중복 관리

output/articles_daily.xlsx 파일에 기사 데이터를 누적 저장한다.
시트 구성: raw_articles, run_log, error_log
"""

import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import List, Set, Optional
from dataclasses import asdict

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import OUTPUT_FILEPATH, OUTPUT_DIR, now_kst, setup_logger

logger = setup_logger("excel")

# ──────────────────────────────────────────────
# 시트별 컬럼 정의
# ──────────────────────────────────────────────
RAW_COLUMNS = [
    ("article_id", "기사ID", 20),
    ("title", "제목", 60),
    ("organization", "기관명", 25),
    ("publish_date", "게시일", 15),
    ("detail_url", "상세URL", 50),
    ("body_text", "본문", 80),
    ("collected_at", "수집일시", 22),
    ("keywords", "키워드", 40),
    ("ai_relevance", "AI관련도", 12),
    ("summary", "요약", 60),
    ("category", "카테고리", 20),
]

RUN_LOG_COLUMNS = [
    ("run_time", "실행일시", 22),
    ("total_found", "발견건수", 12),
    ("new_count", "신규건수", 12),
    ("skipped_count", "중복건수", 12),
    ("error_count", "실패건수", 12),
    ("elapsed_seconds", "소요시간(초)", 15),
]

ERROR_LOG_COLUMNS = [
    ("url", "실패URL", 50),
    ("title", "제목", 40),
    ("error", "에러메시지", 60),
    ("timestamp", "발생시각", 22),
]

# 스타일 상수
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class ExcelManager:
    """엑셀 파일 관리 클래스"""

    def __init__(self, filepath: Optional[Path] = None):
        self.filepath = filepath or OUTPUT_FILEPATH
        self.filepath.parent.mkdir(parents=True, exist_ok=True)

    def _ensure_workbook(self) -> Workbook:
        """워크북이 존재하면 로드, 없으면 생성한다."""
        if self.filepath.exists():
            try:
                return load_workbook(str(self.filepath))
            except Exception as e:
                logger.error(f"엑셀 파일 로드 실패: {e}")
                # 백업 후 재생성
                backup = self.filepath.with_suffix(".xlsx.bak")
                shutil.copy2(self.filepath, backup)
                logger.info(f"기존 파일 백업: {backup}")

        return self._create_workbook()

    def _create_workbook(self) -> Workbook:
        """새 워크북을 생성하고 시트/헤더를 설정한다."""
        wb = Workbook()

        # raw_articles 시트
        ws_raw = wb.active
        ws_raw.title = "raw_articles"
        self._setup_sheet(ws_raw, RAW_COLUMNS)

        # run_log 시트
        ws_run = wb.create_sheet("run_log")
        self._setup_sheet(ws_run, RUN_LOG_COLUMNS)

        # error_log 시트
        ws_err = wb.create_sheet("error_log")
        self._setup_sheet(ws_err, ERROR_LOG_COLUMNS)

        logger.info(f"새 엑셀 파일 생성: {self.filepath}")
        return wb

    def _setup_sheet(self, ws, columns: list) -> None:
        """시트에 헤더를 설정하고 서식을 적용한다."""
        headers = [col[1] for col in columns]
        widths = [col[2] for col in columns]

        # 헤더 쓰기
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # 컬럼 너비 설정
        for col_idx, width in enumerate(widths, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        # 헤더 행 고정
        ws.freeze_panes = "A2"

        # 자동 필터 적용
        if headers:
            last_col = get_column_letter(len(headers))
            ws.auto_filter.ref = f"A1:{last_col}1"

    def get_existing_article_ids(self) -> Set[str]:
        """이미 저장된 기사 ID 집합을 반환한다 (중복 방지용)."""
        if not self.filepath.exists():
            return set()

        try:
            df = pd.read_excel(
                str(self.filepath),
                sheet_name="raw_articles",
                usecols=["기사ID"],
                engine="openpyxl",
            )
            ids = set(df["기사ID"].dropna().astype(str).tolist())
            logger.info(f"기존 기사 {len(ids)}건 로드 (중복 확인용)")
            return ids
        except Exception as e:
            logger.warning(f"기존 기사 ID 로드 실패: {e}")
            return set()

    def save_articles(self, articles: list) -> int:
        """기사 데이터를 raw_articles 시트에 추가한다.

        Returns:
            실제 저장된 건수
        """
        if not articles:
            logger.info("저장할 기사가 없습니다.")
            return 0

        try:
            wb = self._ensure_workbook()
            ws = wb["raw_articles"]

            saved = 0
            for article in articles:
                # Article 객체 또는 dict 처리
                if hasattr(article, "__dict__"):
                    data = {k: v for k, v in article.__dict__.items()}
                else:
                    data = article

                row_data = []
                for col_key, _, _ in RAW_COLUMNS:
                    value = data.get(col_key, "")
                    # 본문이 너무 길면 자르기 (엑셀 셀 제한: 32,767자)
                    if col_key == "body_text" and isinstance(value, str) and len(value) > 30000:
                        value = value[:30000] + "... (잘림)"
                    row_data.append(value)

                ws.append(row_data)
                saved += 1

            # 데이터 행에 테두리 적용
            for row in ws.iter_rows(
                min_row=ws.max_row - saved + 1,
                max_row=ws.max_row,
                min_col=1,
                max_col=len(RAW_COLUMNS),
            ):
                for cell in row:
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            self._safe_save(wb)
            logger.info(f"기사 {saved}건 저장 완료")
            return saved

        except PermissionError:
            logger.error(
                "엑셀 파일이 다른 프로그램에서 열려 있습니다. "
                "파일을 닫고 다시 시도하세요."
            )
            return self._save_to_temp(articles)
        except Exception as e:
            logger.error(f"기사 저장 실패: {e}")
            return 0

    def save_run_log(
        self,
        run_time: str,
        total_found: int,
        new_count: int,
        skipped_count: int = 0,
        error_count: int = 0,
        elapsed_seconds: float = 0.0,
    ) -> None:
        """실행 이력을 run_log 시트에 기록한다."""
        try:
            wb = self._ensure_workbook()
            ws = wb["run_log"]

            ws.append([
                run_time,
                total_found,
                new_count,
                skipped_count,
                error_count,
                elapsed_seconds,
            ])

            self._safe_save(wb)
            logger.info(f"실행 로그 기록 완료 (신규 {new_count}건)")

        except Exception as e:
            logger.error(f"실행 로그 기록 실패: {e}")

    def save_error_log(self, errors: List[dict]) -> None:
        """에러 정보를 error_log 시트에 기록한다."""
        if not errors:
            return

        try:
            wb = self._ensure_workbook()
            ws = wb["error_log"]

            for err in errors:
                ws.append([
                    err.get("url", ""),
                    err.get("title", ""),
                    err.get("error", ""),
                    err.get("timestamp", ""),
                ])

            self._safe_save(wb)
            logger.info(f"에러 로그 {len(errors)}건 기록")

        except Exception as e:
            logger.error(f"에러 로그 기록 실패: {e}")

    def get_recent_articles(self, limit: int = 20) -> List[dict]:
        """최근 수집된 기사 목록을 반환한다 (대시보드용)."""
        if not self.filepath.exists():
            return []

        try:
            df = pd.read_excel(
                str(self.filepath),
                sheet_name="raw_articles",
                engine="openpyxl",
            )
            if df.empty:
                return []

            # 최근 N건
            df = df.tail(limit).iloc[::-1]  # 역순 (최신 순)

            # 본문은 미리보기만
            if "본문" in df.columns:
                df["본문"] = df["본문"].astype(str).str[:200] + "..."

            return df.to_dict("records")
        except Exception as e:
            logger.warning(f"최근 기사 조회 실패: {e}")
            return []

    def get_filtered_articles(self, start_date: str = None, end_date: str = None, limit: int = 100) -> List[dict]:
        """날짜 범위로 필터링된 기사 목록을 반환한다."""
        if not self.filepath.exists():
            return []

        try:
            df = pd.read_excel(
                str(self.filepath),
                sheet_name="raw_articles",
                engine="openpyxl",
            )
            if df.empty:
                return []

            # 필터링 적용
            if "게시일" in df.columns:
                df["게시일"] = df["게시일"].astype(str)
                if start_date:
                    df = df[df["게시일"] >= start_date]
                if end_date:
                    df = df[df["게시일"] <= end_date]

            # 최신 순 정렬 (수집일시 기준 또는 역순)
            df = df.iloc[::-1]

            if limit:
                df = df.head(limit)

            # 본문은 미리보기만
            if "본문" in df.columns:
                df["본문"] = df["본문"].astype(str).str[:200] + "..."

            # NaN 처리
            df = df.fillna("")

            return df.to_dict("records")
        except Exception as e:
            logger.warning(f"필터링 기사 조회 실패: {e}")
            return []

    def create_temp_filtered_excel(self, start_date: str, end_date: str) -> Optional[Path]:
        """지정 기간의 데이터를 필터링하여 임시 엑셀 파일을 생성하고 경로를 반환한다."""
        if not self.filepath.exists():
            return None

        try:
            df = pd.read_excel(
                str(self.filepath),
                sheet_name="raw_articles",
                engine="openpyxl",
            )
            if df.empty:
                return None

            # 필터링 적용
            if "게시일" in df.columns:
                df["게시일"] = df["게시일"].astype(str)
                if start_date:
                    df = df[df["게시일"] >= start_date]
                if end_date:
                    df = df[df["게시일"] <= end_date]

            # 임시 파일 경로
            temp_filename = f"articles_filtered_{start_date}_{end_date}.xlsx"
            temp_path = self.filepath.parent / temp_filename

            # openpyxl로 새 엑셀 만들기 및 스타일링
            wb = Workbook()
            ws = wb.active
            ws.title = "raw_articles"
            self._setup_sheet(ws, RAW_COLUMNS)

            # 데이터 추가
            for _, row in df.iterrows():
                row_data = []
                for col_key, _, _ in RAW_COLUMNS:
                    # df의 컬럼명은 한글이므로 한글 헤더명을 찾아 매핑해야 함
                    hangul_col = next((c[1] for c in RAW_COLUMNS if c[0] == col_key), None)
                    value = row.get(hangul_col, "")
                    if pd.isna(value):
                        value = ""
                    row_data.append(value)
                ws.append(row_data)

            # 스타일 입히기
            for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(RAW_COLUMNS)):
                for cell in r:
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            wb.save(str(temp_path))
            logger.info(f"임시 필터링 엑셀 생성 완료: {temp_path}")
            return temp_path

        except Exception as e:
            logger.error(f"임시 필터링 엑셀 생성 실패: {e}")
            return None

    def get_run_logs(self, limit: int = 10) -> List[dict]:
        """최근 실행 로그를 반환한다."""
        if not self.filepath.exists():
            return []

        try:
            df = pd.read_excel(
                str(self.filepath),
                sheet_name="run_log",
                engine="openpyxl",
            )
            if df.empty:
                return []

            return df.tail(limit).iloc[::-1].to_dict("records")
        except Exception as e:
            logger.warning(f"실행 로그 조회 실패: {e}")
            return []

    def get_error_logs(self, limit: int = 20) -> List[dict]:
        """최근 에러 로그를 반환한다."""
        if not self.filepath.exists():
            return []

        try:
            df = pd.read_excel(
                str(self.filepath),
                sheet_name="error_log",
                engine="openpyxl",
            )
            if df.empty:
                return []

            return df.tail(limit).iloc[::-1].to_dict("records")
        except Exception as e:
            logger.warning(f"에러 로그 조회 실패: {e}")
            return []

    def get_today_count(self) -> int:
        """오늘 수집된 기사 수를 반환한다."""
        if not self.filepath.exists():
            return 0

        try:
            df = pd.read_excel(
                str(self.filepath),
                sheet_name="raw_articles",
                usecols=["수집일시"],
                engine="openpyxl",
            )
            if df.empty:
                return 0

            today = now_kst().strftime("%Y-%m-%d")
            df["수집일시"] = df["수집일시"].astype(str)
            return int(df["수집일시"].str.startswith(today).sum())
        except Exception as e:
            logger.warning(f"오늘 수집 건수 조회 실패: {e}")
            return 0

    def get_last_run_time(self) -> Optional[str]:
        """마지막 실행 시간을 반환한다."""
        logs = self.get_run_logs(limit=1)
        if logs:
            return str(logs[0].get("실행일시", ""))
        return None

    def _safe_save(self, wb: Workbook) -> None:
        """안전하게 워크북을 저장한다 (파일 잠금 대응)."""
        try:
            wb.save(str(self.filepath))
        except PermissionError:
            # 임시 파일로 저장
            temp_path = self.filepath.with_suffix(".xlsx.tmp")
            wb.save(str(temp_path))
            logger.warning(
                f"원본 파일 잠금 → 임시 파일로 저장: {temp_path}\n"
                "원본 파일을 닫은 후 수동으로 교체하세요."
            )

    def _save_to_temp(self, articles: list) -> int:
        """파일 잠금 시 임시 파일로 저장한다."""
        try:
            temp_path = self.filepath.with_name(
                f"articles_temp_{now_kst().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            temp_manager = ExcelManager(temp_path)
            result = temp_manager.save_articles(articles)
            logger.warning(f"임시 파일로 {result}건 저장: {temp_path}")
            return result
        except Exception as e:
            logger.error(f"임시 파일 저장도 실패: {e}")
            return 0
